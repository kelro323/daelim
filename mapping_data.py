from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import math
import time


# Category 가운데 정렬 및 진하게
def bold(worksheet, cell_name):
    cells = worksheet[cell_name]
    cells.font = Font(bold=True)
    cells.alignment = Alignment(horizontal='center', vertical='center')


# excel file array 로 변환
def excel_to_array(load_ws):
    values = []
    for row in load_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        values.append(row_value)
    return values


# 같은 시간대 데이터 하나만 남기기 위해서
# 동일 시간대 데이터는 뒤쪽 데이터를 남김
def remove_same_time(all_value):
    result_list = []
    for i in range(0, len(all_value) - 1):
        if all_value[i][0] == all_value[i + 1][0]:
            continue
        else:
            result_list.append(all_value[i])
    result_list.append(all_value[len(all_value)-1])
    return result_list


start_time = time.time()


# 파일위치 하드코딩
# 해당 엑셀 파일은 카테고리 없이 1열부터 데이터 시작
load_wb_ocr = load_workbook("C:\\Users\\Seungwoo\\Documents\\대림산업\\190627_ocr.xlsx", data_only=True)
load_wb_swing = load_workbook("C:\\Users\\Seungwoo\\Documents\\대림산업\\190627_swing.xlsx", data_only=True)


# 혹시 Sheet 이름이 다르면 시트 이름 변경
load_ws_ocr = load_wb_ocr['Sheet1']
all_value_ocr = remove_same_time(excel_to_array(load_ws_ocr))

load_ws_swing = load_wb_swing['Sheet1']
all_value_swing = remove_same_time(excel_to_array(load_ws_swing))

ocr_index = 0
for value in all_value_swing:
    print(value)
    find_date = False
    value_date = value[0]
    for i in range(ocr_index, len(all_value_ocr)):
        try:
            find_index = all_value_ocr[i].index(value_date)
            if find_index == 0:
                find_date = True
                value.append(all_value_ocr[i][1])
                value.append(all_value_ocr[i][2])
                x_coord = -math.sin(math.radians(value[1])) * value[2]
                y_coord = math.cos(math.radians(value[1])) * value[2]
                value.append(x_coord)
                value.append(y_coord)
                ocr_index = i
                break
        except ValueError:
            continue
    if not find_date:
        value.append('none')
        value.append('none')

write_wb = Workbook()
write_ws = write_wb.active
write_ws.append(['datetime', 'yaw', 'trolley_length', 'weight', 'x_coord', 'y_coord'])
cell_list = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']
for c in cell_list:
    bold(write_ws, c)
for i in range(0, len(all_value_swing)):
    write_ws.append(all_value_swing[i])

# 파일 출력 경로
write_wb.save("C:\\Users\\Seungwoo\\Documents\\대림산업\\190627_daily.xlsx")
print(time.time()-start_time)
